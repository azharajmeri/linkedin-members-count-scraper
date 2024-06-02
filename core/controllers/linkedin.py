import os
import time

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

from config import OUTPUT_FILE_PATH, INPUT_FILE_PATH, LINKEDIN_EMAIL, LINKEDIN_PASSWORD
from core.exceptions.web_driver import UnableToLoginException
from core.handler.web_driver_tabs import TabHandler
from core.managers.web_driver import WebDriverManager


class LinkedInController:
    def __init__(self):
        self.driver_manager = WebDriverManager()
        self.driver_manager.start_driver()
        self.companies = []
        self.excel_file = OUTPUT_FILE_PATH

    def initiate(self):
        print("INITIATING SCRAPER")
        self.driver_manager.navigate_to("https://www.linkedin.com/login", raise_exception=True)
        self.read_excel()
        self.initiate_scraping()

    def read_excel(self):
        wb = openpyxl.load_workbook(INPUT_FILE_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
            for cell in row:
                if cell:
                    self.companies.append(cell)

    def initiate_scraping(self):
        try:
            self.check_login()
        except Exception:
            raise UnableToLoginException
        if self.driver_manager.driver.current_url == "https://www.linkedin.com/feed/":
            self.extract_members_details()
        else:
            raise UnableToLoginException(f"Unable to login - {self.driver_manager.driver.current_url}")

    def check_login(self):
        username = password = None
        try:
            username = self.driver_manager.find_element((By.XPATH, '//input[@id="username"]'))
        except NoSuchElementException:
            print("Username field not available on page.")

        try:
            password = self.driver_manager.find_element((By.XPATH, '//input[@id="password"]'))
        except NoSuchElementException:
            print("password field not available on page.")

        if username and password:
            username.clear()
            username.send_keys(LINKEDIN_EMAIL)

            password.clear()
            password.send_keys(LINKEDIN_PASSWORD)
            password.send_keys(Keys.ENTER)

        password = None
        try:
            password = self.driver_manager.find_element((By.XPATH, '//input[@id="password"]'))
        except NoSuchElementException:
            print("password field not available on page.")

        if password:
            password.clear()
            password.send_keys(LINKEDIN_PASSWORD)
            password.send_keys(Keys.ENTER)

    def extract_members_details(self):
        # all_data = []
        self.initialize_excel()
        self._extract_members_details()
        # all_data.append(data)
        # self._save_to_excel(all_data)

    def initialize_excel(self):
        # Check if the file exists
        file_exists = os.path.exists(self.excel_file)

        if file_exists:
            # Load the existing workbook
            self.wb = load_workbook(self.excel_file)
            self.ws = self.wb.active
            # Clear existing data
            self.ws.delete_rows(1, self.ws.max_row)
        else:
            # Create a new workbook and select the active sheet
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Members Details"

        # Define headers
        # headers = ["Company", "Total Members", "Top Country 1", "Top Country 2", "Top Country 3", "Top Country 4",
        #            "Top Country 5", "Top Country 6", "Top Country 7", "Top Country 8", "Top Country 9",
        #            "Top Country 10"]
        headers = ["Company", "Total Members"]
        self.ws.append(headers)


        # self.ws.append(data)

        self.wb.save(self.excel_file)

        # print(f"Data saved successfully to {self.excel_file}")

    def _extract_members_details(self):
        tab_handler = TabHandler(
            self.driver_manager,
            (By.XPATH, '//h2[text()[contains(.,"associated members")]]'),
            self.ws,
            self.wb,
            self.excel_file
        )
        start = time.time()
        tab_handler.open_urls_in_tabs(self.companies)
        print(f"TOTAL RUNTIME OF THE PROGRAM: {time.time() - start}")
